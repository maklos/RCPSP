import time
import os
import glob
from itertools import chain, combinations
import random
import statistics as st
import re
import multiprocessing as mp
import tensorflow as tf
import tflearn
from tflearn.layers.core import input_data, dropout, fully_connected
from tflearn.optimizers import SGD
from tflearn.layers.estimator import regression
from datetime import datetime
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side



#TODO
#test CPU parallelisation -> it is working but it stops sometimes when using pool.map
#test GPU parallelization -> probably not possible because the GPU do not accept parallel tasks
#tensorflow GPU with float 16bit
#run generate data and train in parallel avoiding that the buffer becomes too big
#generate extra activity sequences
#assign priority values also to ready to start activities outside the state vector
#introduce check for existing neural network in path
#test different learning parameters
#test with stochastic times and luck factors
#global (SL) + local (RL)
#separate codes in different scripts if the code becomes too big (main, runSimulation)
#try discount factor for resource utilization in the future
#also test value network
#test different learning libraries (tf, tflearn, keras, ...)

#MAYBE TO BE MODIFIED
#take same number of random files from each class
#tools for better visualization (GUI simulation)
#decision policy based on "prefer activities with a lot of resource of a certain type required"



######################################## FUNCTIONS ########################################
def runSimulation(runSimulation_input):
    currentActivitySequence = runSimulation_input.activitySequence
    numberOfSimulationRuns = runSimulation_input.numberOfSimulationRuns
    timeDistribution = runSimulation_input.timeDistribution
    purpose = runSimulation_input.purpose
    randomDecisionProbability = runSimulation_input.randomDecisionProbability
    policyType = runSimulation_input.policyType
    decisionTool = runSimulation_input.decisionTool
    numberOfResources = runSimulation_input.numberOfResources
    numberOfActivitiesInStateVector = runSimulation_input.numberOfActivitiesInStateVector
    stateVectorLength = runSimulation_input.stateVectorLength
    decisions_indexActivity = runSimulation_input.decisions_indexActivity
    rescaleFactorTime = runSimulation_input.rescaleFactorTime
    numberOfActivities = runSimulation_input.numberOfActivities

    print("start " + str(currentActivitySequence.fileName[:-4]))
    # print('------------------------------------------------------------------------------------------')

    # reset variables for the series of runs
    indexSimulationRun = 0
    # reset lists for the series of runs
    totalDurations = []
    luckFactors = []
    trivialDecisionPercentages = []
    stateActionPairsOfRuns = []

    while indexSimulationRun < numberOfSimulationRuns:
        # reset variables for run
        sumActivityDuration = 0
        step = 0
        numberOfDecisions = 0
        numberOfTrivialDecisions = 0

        # reset lists for run
        if purpose == "generateData":
            currentStateActionPairsOfRun = []

        # reset dynamic variables of classes for run
        currentActivitySequence.availableResources = currentActivitySequence.totalResources[:]
        currentActivitySequence.virtualTime = 0
        for i in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[i].withToken = False
            currentActivitySequence.activities[i].idleToken = False
            currentActivitySequence.activities[i].numberOfCompletedPreviousActivities = 0
            currentActivitySequence.activities[i].remainingTime = 0
            currentActivitySequence.activities[i].processedTime = 0
            currentActivitySequence.activities[i].seizedResources = [0] * numberOfResources

        # set startActivities ready to start
        for indexStartActivity in currentActivitySequence.indexStartActivities:
            currentActivitySequence.activities[indexStartActivity].withToken = True
            currentActivitySequence.activities[indexStartActivity].idleToken = True

        # start simulation
        simulationRunFinished = False
        while simulationRunFinished == False:  # if there are some token left in some activities
            step += 1

            ## STEP 1 ##
            # 1.1 find activities ready to start
            indexReadyToStartActivities = []
            for i, currentActivity in enumerate(currentActivitySequence.activities):
                if (
                        currentActivity.withToken and currentActivity.idleToken and currentActivity.numberOfCompletedPreviousActivities == currentActivity.numberOfPreviousActivities):
                    # verify that enough resources are available for every ready to start activity
                    enoughResourcesAreAvailable = True
                    for j in range(numberOfResources):
                        if currentActivity.requiredResources[j] > currentActivitySequence.availableResources[j]:
                            enoughResourcesAreAvailable = False
                            break
                    if enoughResourcesAreAvailable:
                        indexReadyToStartActivities.append(i)
            # print('indexReadyToStartActivities',indexReadyToStartActivities)

            # 1.2 check if the decision is trivial
            trivialDecision = True
            indexReadyToStartActivitiesInState = indexReadyToStartActivities[0:min(numberOfActivitiesInStateVector,
                                                                                   len(indexReadyToStartActivities))]
            # compute powerset of decisions_indexActivity
            indexReadyToStartActivitiesPowerset = list(powerset(indexReadyToStartActivitiesInState))
            # find feasible combined decisions_indexActivity (only resource check needed)
            feasibleCombinedDecisions_indexActivity = []
            for i in reversed(range(len(indexReadyToStartActivitiesPowerset))):
                currentDecision = list(indexReadyToStartActivitiesPowerset[i])
                decisionIsASubsetOfFeasibleDecision = False
                for j, feasibleDecisionAlreadyInList in enumerate(feasibleCombinedDecisions_indexActivity):
                    if len(set(currentDecision) - set(feasibleDecisionAlreadyInList)) == 0:
                        decisionIsASubsetOfFeasibleDecision = True
                        break
                if decisionIsASubsetOfFeasibleDecision == False:
                    # verify that enough resources are available to start all the activities
                    totalRequiredResources = [0] * numberOfResources
                    for indexCurrentActivity in currentDecision:
                        for j in range(numberOfResources):
                            totalRequiredResources[j] += \
                            currentActivitySequence.activities[indexCurrentActivity].requiredResources[j]
                    enoughResourcesAreAvailable = True
                    for j in range(numberOfResources):
                        if totalRequiredResources[j] > currentActivitySequence.availableResources[j]:
                            enoughResourcesAreAvailable = False
                            break
                    if enoughResourcesAreAvailable:
                        feasibleCombinedDecisions_indexActivity.append(currentDecision)
            if len(feasibleCombinedDecisions_indexActivity) > 1:
                trivialDecision = False
                # print('not trivialDecision')

            numberOfDecisions += 1
            if trivialDecision:
                numberOfTrivialDecisions += 1
                # print('numberOfTrivialDecisions',numberOfTrivialDecisions)

            # 1.3 define activity conversion vector and resource conversion vector
            # initialise activityConversionVector and ResourceConversionVector
            activityConversionVector = [-1] * numberOfActivitiesInStateVector
            activityScores = []
            indexReadyToStartActivitiesInState = indexReadyToStartActivities[0:min(numberOfActivitiesInStateVector,
                                                                                   len(indexReadyToStartActivities))]
            # print('indexReadyToStartActivitiesInState',indexReadyToStartActivitiesInState)

            if trivialDecision:
                # no conversion needed
                resourceConversionVector = list(range(0, numberOfResources))
                for i in range(len(indexReadyToStartActivitiesInState)):
                    activityConversionVector[i] = indexReadyToStartActivitiesInState[i]
            else:
                # conversion is required
                # find most critical resources (i.e. required resources to start the ready to start activities normalized by the total resources)
                resourceNeedForReadyToStartActivities = [0] * numberOfResources
                for i in indexReadyToStartActivities:
                    for j in range(numberOfResources):
                        resourceNeedForReadyToStartActivities[j] += \
                        currentActivitySequence.activities[i].requiredResources[j] / \
                        currentActivitySequence.totalResources[j]
                        # print('resourceNeedForReadyToStartActivities',resourceNeedForReadyToStartActivities)

                # create resourceConversionVector
                indexResourcesGlobal = list(range(0, numberOfResources))
                indexResourcesGlobal_reordered = [x for _, x in sorted(
                    zip(resourceNeedForReadyToStartActivities, indexResourcesGlobal), reverse=True)]
                resourceConversionVector = indexResourcesGlobal_reordered
                # print('resourceConversionVector',resourceConversionVector)

                # reorder activities depending on resource utilisation
                activityScores = [-1] * numberOfActivitiesInStateVector

                for i in range(len(indexReadyToStartActivitiesInState)):
                    for j in range(len(resourceConversionVector)):
                        resourceMultiplicator = 100 ** (numberOfResources - j - 1)
                        # print('resourceMultiplicator',resourceMultiplicator)
                        resourceQuantity = \
                        currentActivitySequence.activities[indexReadyToStartActivitiesInState[i]].requiredResources[
                            resourceConversionVector[j]]
                        activityScores[i] += 1 + resourceMultiplicator * resourceQuantity
                        # print('activityScores',activityScores)

                indexActivitiesGlobal = [-1] * numberOfActivitiesInStateVector
                indexActivitiesGlobal[0:len(indexReadyToStartActivitiesInState)] = indexReadyToStartActivitiesInState
                indexActivitiesGlobal_reordered = [x for _, x in
                                                   sorted(zip(activityScores, indexActivitiesGlobal), reverse=True)]
                activityConversionVector = indexActivitiesGlobal_reordered
                # print('activityConversionVector',activityConversionVector)

            # 1.4 normalized state vector and matrix are created
            currentState_readyToStartActivities = []

            if trivialDecision == False:
                currentState_readyToStartActivities = np.zeros([stateVectorLength])

                for i, indexActivity in enumerate(activityConversionVector):

                    if indexActivity != -1:
                        currentState_readyToStartActivities[0 + i * (1 + numberOfResources)] = \
                        currentActivitySequence.activities[indexActivity].time * rescaleFactorTime

                        for j in range(numberOfResources):
                            currentState_readyToStartActivities[1 + j + i * (1 + numberOfResources)] = \
                            currentActivitySequence.activities[indexActivity].requiredResources[
                                resourceConversionVector[j]] / currentActivitySequence.totalResources[
                                resourceConversionVector[j]]

                for j in range(numberOfResources):
                    currentState_readyToStartActivities[
                        numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + j] = \
                    currentActivitySequence.availableResources[resourceConversionVector[j]] / \
                    currentActivitySequence.totalResources[resourceConversionVector[j]]
            # (optional: add information about the future resource utilisation)
            # determine the earliest starting point of each activity considering the problem without resource constraints and deterministic
            # currentState_futureResourceUtilisation = np.zeros([numberOfResources, timeHorizon])

            # 1.5 Use the policy and the decision tool to define which tokens can begin the correspondent activity or remain idle
            randomDecisionAtThisStep = (random.random() < randomDecisionProbability)
            # print('randomDecisionAtThisStep',randomDecisionAtThisStep)

            if trivialDecision:  # if the decision is trivial, it does not matter how the priority values are assigned
                randomDecisionAtThisStep = True

            if randomDecisionAtThisStep:
                priorityValues = np.random.rand(numberOfActivitiesInStateVector)
                # print('randomDecisionAtThisStep')

            else:
                if policyType == "neuralNetworkModel":
                    currentState_readyToStartActivities = currentState_readyToStartActivities.reshape(-1,
                                                                                                      stateVectorLength,
                                                                                                      1)
                    # print('currentState_readyToStartActivities:',currentState_readyToStartActivities)
                    outputNeuralNetworkModel = decisionTool.predict(currentState_readyToStartActivities)
                    # print('outputNeuralNetworkModel:',outputNeuralNetworkModel)
                    priorityValues = outputNeuralNetworkModel[0]
                    # print('priorityValues:', priorityValues)

                elif policyType == "heuristic":
                    # print("generate priority values with most critical resource")
                    priorityValues = [1, 0.8, 0.6, 0.4, 0.2, 0]

                elif policyType == "otherPolicy2":
                    print("generate priority values with other policy 2")
                else:
                    print("policy name not existing")

            # reorder list according to priority
            decisions_indexActivity_reordered = [x for _, x in
                                                 sorted(zip(priorityValues, decisions_indexActivity), reverse=True)]
            # print('decisions_indexActivity_reordered)',decisions_indexActivity_reordered)

            # if not randomDecisionAtThisStep:
            #     print('not randomdecision##############################################################################################')
            #     print('decisions_indexActivity_reordered',decisions_indexActivity_reordered)

            # use the priority values to start new activities
            currentAction = np.zeros([numberOfActivitiesInStateVector])
            indexStartedActivities = []
            # consider the decision one by one in reordered list
            for indexActivityToStartLocal in decisions_indexActivity_reordered:
                indexActivityToStartGlobal = activityConversionVector[indexActivityToStartLocal]
                # print('indexActivityToStartGlobal',indexActivityToStartGlobal)

                if indexActivityToStartGlobal != -1:
                    currentActivity = currentActivitySequence.activities[indexActivityToStartGlobal]

                    if currentActivity.withToken and currentActivity.idleToken and currentActivity.numberOfCompletedPreviousActivities == currentActivity.numberOfPreviousActivities:
                        # verify that enough resources are available to start
                        enoughResourcesAreAvailable = True

                        for i in range(numberOfResources):
                            if currentActivity.requiredResources[i] > currentActivitySequence.availableResources[i]:
                                enoughResourcesAreAvailable = False
                                break

                        if enoughResourcesAreAvailable:
                            currentActivitySequence.activities[indexActivityToStartGlobal].idleToken = False

                            # 1.6 Set remaining time for the starting activity
                            if timeDistribution == "deterministic":
                                currentActivitySequence.activities[indexActivityToStartGlobal].remainingTime = \
                                currentActivitySequence.activities[indexActivityToStartGlobal].time
                                sumActivityDuration += currentActivitySequence.activities[
                                    indexActivityToStartGlobal].remainingTime

                            # 1.7 seize resources
                            for i in range(numberOfResources):
                                currentActivitySequence.activities[indexActivityToStartGlobal].seizedResources[i] = \
                                currentActivitySequence.activities[indexActivityToStartGlobal].requiredResources[i]
                                currentActivitySequence.availableResources[i] -= \
                                currentActivitySequence.activities[indexActivityToStartGlobal].requiredResources[i]

                            # update the action vector with the activity that has been just started
                            currentAction[indexActivityToStartLocal] = 1
                            indexStartedActivities.append(indexActivityToStartGlobal)

            # print('currentaction',currentAction)

            # 1.8 if the purpose is to generate training data, save the current state action pair
            if purpose == "generateData" and trivialDecision == False:
                currentStateActionPair = stateActionPair()
                currentStateActionPair.state = currentState_readyToStartActivities
                # print('currentState_readyToStartActivities',currentState_readyToStartActivities)
                currentStateActionPair.action = currentAction
                # print('currentAction',currentAction)
                currentStateActionPairsOfRun.append(currentStateActionPair)

            ## STEP 2 ##
            # 2.1 find out when the next event (activity end) occurs
            smallestRemainingTime = 1e300
            indexActiveActivities = []
            for i in range(numberOfActivities):
                if currentActivitySequence.activities[i].withToken and currentActivitySequence.activities[
                    i].idleToken == False:
                    indexActiveActivities.append(i)
                    if currentActivitySequence.activities[i].remainingTime < smallestRemainingTime:
                        smallestRemainingTime = currentActivitySequence.activities[i].remainingTime

            # 2.2 find next finishing activities
            indexNextFinishingActivities = []
            for i in indexActiveActivities:
                if currentActivitySequence.activities[i].remainingTime == smallestRemainingTime:
                    indexNextFinishingActivities.append(i)

            # 2.3 jump forward to activity end
            currentActivitySequence.virtualTime += smallestRemainingTime
            for i in indexActiveActivities:
                currentActivitySequence.activities[i].remainingTime -= smallestRemainingTime
                currentActivitySequence.activities[i].processedTime += smallestRemainingTime

            ## STEP 3 ##
            # for each just finished activity:
            for i in indexNextFinishingActivities:
                # 3.1 find following activities
                indexFollowingActivities = currentActivitySequence.activities[i].indexFollowingActivities
                # 3.2 for each following activity, increment the numberOfCompletedPreviousActivities and, if there is no token already in the following activity, add an idle token.
                for j in indexFollowingActivities:
                    currentActivitySequence.activities[j].numberOfCompletedPreviousActivities += 1
                    if currentActivitySequence.activities[j].withToken == False:
                        currentActivitySequence.activities[j].withToken = True
                        currentActivitySequence.activities[j].idleToken = True
                # 3.3 delete token from just finished activity
                currentActivitySequence.activities[i].withToken = False
                currentActivitySequence.activities[i].idleToken = False
                # 3.4 release resources back to the resource pool
                currentActivitySequence.activities[i].seizedResources = [0] * numberOfResources
                for j in range(numberOfResources):
                    currentActivitySequence.availableResources[j] += \
                    currentActivitySequence.activities[i].requiredResources[j]

            ## STEP 4 ##
            # check if all activities are completed (i.e. no more token)
            simulationRunFinished = True
            for i in range(numberOfActivities):
                if currentActivitySequence.activities[i].withToken:
                    simulationRunFinished = False
                    break

        totalDuration = currentActivitySequence.virtualTime
        luckFactor = sumActivityDuration / sum(a.time for a in currentActivitySequence.activities)
        trivialDecisionPercentage = numberOfTrivialDecisions / numberOfDecisions

        totalDurations.append(totalDuration)
        luckFactors.append(luckFactor)
        trivialDecisionPercentages.append(trivialDecisionPercentage)

        if purpose == "generateData":
            stateActionPairsOfRuns.append(currentStateActionPairsOfRun)
        # print('trivialDecisionPercentages',trivialDecisionPercentages)

        # increment the index for the simulation run at the end of the loop
        indexSimulationRun += 1

    totalDurationMean = st.mean(totalDurations)
    totalDurationStDev = None
    if numberOfSimulationRuns != 1:
        totalDurationStDev = st.stdev(totalDurations)
    totalDurationMin = min(totalDurations)
    totalDurationMax = max(totalDurations)
    luckFactorMean = st.mean(luckFactors)
    trivialDecisionPercentageMean = st.mean(trivialDecisionPercentages)

    currentRunSimulation_output = runSimulation_output()
    currentRunSimulation_output.totalDurationMean = totalDurationMean
    currentRunSimulation_output.totalDurationStDev = totalDurationStDev
    currentRunSimulation_output.totalDurationMin = totalDurationMin
    currentRunSimulation_output.totalDurationMax = totalDurationMax
    currentRunSimulation_output.luckFactorMean = luckFactorMean
    currentRunSimulation_output.trivialDecisionPercentageMean = trivialDecisionPercentageMean

    # submit the stateActionPairs of the best run, if the standard deviation of the duration is not zero
    if purpose == "generateData":
        if totalDurationStDev != 0:
            indexBestRun = totalDurations.index(totalDurationMin)
            currentRunSimulation_output.stateActionPairsOfBestRun = stateActionPairsOfRuns[indexBestRun]

    print("end " + str(currentActivitySequence.fileName[:-4]))
    # print('-------------------------------------------------------------')

    return currentRunSimulation_output


def createNeuralNetworkModel(input_size, output_size, learningRate):
    network = input_data(shape=[None, input_size, 1], name="input")

    tflearn.init_graph(num_cores=1, gpu_memory_fraction=0.2)

    network = fully_connected(network, 12, activation="sigmoid", regularizer='L2',
                              weights_init=tf.constant_initializer(0.03))

    network = fully_connected(network, output_size, activation="softmax", weights_init=tf.constant_initializer(0.01),
                              regularizer='L2')

    network = regression(network, optimizer='RMSProp', learning_rate=learningRate, loss="categorical_crossentropy",
                         name="targets")

    model = tflearn.DNN(network, tensorboard_dir="log")

    return model


def powerset(listOfElements):
    s = list(listOfElements)
    return chain.from_iterable(combinations(s, r) for r in range(len(s) + 1))


######################################## CLASSES ########################################
class activitySequence:

    def __init__(self):
        # static (do not change during simulation)
        self.index = None
        self.fileName = None
        self.numberOfResources = None
        self.totalResources = []
        self.numberOfActivities = None
        self.activities = []
        self.indexStartActivities = []

        # dynamic (change during simulation)
        self.availableResources = []
        self.totalDurationMean = None
        self.totalDurationStandardDeviation = None
        self.totalDurationMin = None
        self.totalDurationMax = None
        self.luckFactorMean = None
        self.totalDurationWithPolicy = None
        self.totalDurationWithHeuristic = None
        self.trivialDecisionPercentageMean = None


class activity:

    def __init__(self):
        # static (do not change during simulation)
        self.index = None
        self.time = None  # expected value. Only deterministic component. The distribution is given as an argument in the function run simulation.
        self.requiredResources = []
        self.numberOfPreviousActivities = 0
        self.indexFollowingActivities = []

        # dynamic (change during simulation)
        self.withToken = None
        self.idleToken = None
        self.numberOfCompletedPreviousActivities = None
        self.remainingTime = None  # time to activity end
        self.processedTime = None  # time elapsed from the beginning of the activity
        self.seizedResources = []


class stateActionPair:
    def __init__(self):
        self.state = None
        self.action = None


class runSimulation_input:
    def __init__(self):
        self.activitySequence = []
        self.numberOfSimulationRuns = None
        self.timeDistribution = None
        self.purpose = None
        self.randomDecisionProbability = None
        self.policyType = None
        self.decisionTool = None
        self.numberOfResources = None
        self.numberOfActivitiesInStateVector = None
        self.stateVectorLength = None
        self.decisions_indexActivity = []
        self.rescaleFactorTime = None
        self.numberOfActivities = None


class runSimulation_output:
    def __init__(self):
        self.totalDurationMean = None
        self.totalDurationStDev = None
        self.totalDurationMin = None
        self.totalDurationMax = None
        self.luckFactorMean = None
        self.trivialDecisionPercentageMean = None
        self.stateActionPairsOfBestRun = []


######################################## MAIN ########################################
t_start = time.time()

# user defined parameters
# problem parameters
timeDistribution = "deterministic"  # deterministic, exponential, uniform_1, uniform_2, ...

# CPU parameters
numberOfCpuProcessesToGenerateData = 16  # paoloPC has 16 cores
maxTasksPerChildToGenerateData = 4  # 4 is the best for paoloPC

# input state vector  parameters
numberOfActivitiesInStateVector = 6
rescaleFactorTime = 0.1
timeHorizon = 10

# random generation parameters
numberOfSimulationRunsToGenerateData = 1000
numberOfSimulationRunsToTestPolicy = 1
numberOfExperiments = 1

# train parameters
percentageOfFilesTest = 0.1
importExistingNeuralNetworkModel = False
neuralNetworkModelAlreadyExists = False
numberOfEpochs = 1000  # walk entire samples
learningRate = 0.001

# paths
relativePath = os.path.dirname(__file__)
absolutePathProjects = relativePath + "/database/"
# absolutePathProjects = relativePath + "/J30/"
# absolutePathExcelOutput = relativePath + "/Benchmark.xlsx"

# other parameters
np.set_printoptions(precision=4)  # print precision of numpy variables

# initialise variables
numberOfActivities = None
numberOfResources = None
activitySequences = []
decisions_indexActivity = []
decisions_indexActivityPowerset = []
states = []
actions = []
sumTotalDurationRandomTrainRecord = []
sumTotalDurationWithNeuralNetworkModelTrainRecord = []
sumTotalDurationWithHeuristicTrainRecord = []
sumTotalDurationRandomTestRecord = []
sumTotalDurationWithNeuralNetworkModelTestRecord = []
sumTotalDurationWithHeuristicTestRecord = []

# read all activity sequences from database
absolutePathProjectsGlob = absolutePathProjects + "*.txt"
files = sorted(glob.glob(absolutePathProjectsGlob))

# divide all activity sequences in training and test set
numberOfFiles = len(files)
numberOfFilesTest = round(numberOfFiles * percentageOfFilesTest)
numberOfFilesTrain = numberOfFiles - numberOfFilesTest
indexFiles = list(range(0, numberOfFiles))
indexFilesTrain = []
indexFilesTest = []

# choose the first element of every set to test
for i in range(numberOfFilesTest):
    # randomIndex = random.randrange(0, len(indexFiles))
    randomIndex = i * 9
    indexFilesTest.append(indexFiles[randomIndex])
    del indexFiles[randomIndex]  # delete
indexFilesTrain = indexFiles

# organize the read activity sequences in classes
for i in range(numberOfFiles):
    file = files[i]
    # create a new activitySequence object
    currentActivitySequence = activitySequence()
    with open(file, "r") as f:
        currentActivitySequence.index = i
        currentActivitySequence.fileName = os.path.basename(f.name)
        # allLines = f.read()
        # next(f)
        firstLine = f.readline()  # information about numberOfActivities and numberOfResourceTypes
        firstLineDecomposed = re.split(" +", firstLine)
        numberOfActivities = (int(firstLineDecomposed[0]) - 2)  # the first and last dummy activity do not count
        currentActivitySequence.numberOfActivities = numberOfActivities
        secondLine = f.readline()  # information about total available resources
        secondLineDecomposed = re.split(" +", secondLine)
        numberOfResources = 0
        # secondLineDecomposed=[int(secondLineDecomposed)]
        # print(secondLineDecomposed)
        for totalResources in secondLineDecomposed[0:-1]:
            numberOfResources += 1
            currentActivitySequence.totalResources.append(int(totalResources))
        currentActivitySequence.numberOfResources = numberOfResources
        thirdLine = f.readline()  # information about starting activities
        thirdLineDecomposed = re.split(" +", thirdLine)
        for IdActivity in thirdLineDecomposed[6:-1]:
            currentActivitySequence.indexStartActivities.append(int(IdActivity) - 2)
        line = f.readline()
        while line:
            lineDecomposed = re.split(" +", line)
            if int(lineDecomposed[0]) == 0:
                break
            else:
                currentActivity = activity()
                currentActivity.time = int(lineDecomposed[0])
                currentActivity.requiredResources = [int(lineDecomposed[1]), int(lineDecomposed[2]),
                                                     int(lineDecomposed[3]), int(lineDecomposed[4])]
                for IdFollowingActivity in lineDecomposed[6:-1]:
                    if int(
                            IdFollowingActivity) != numberOfActivities + 2:  # if the following action is not the last dummy activity
                        currentActivity.indexFollowingActivities.append(int(IdFollowingActivity) - 2)
            currentActivitySequence.activities.append(currentActivity)
            line = f.readline()
        # add indexes to list of activities
        for j in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[j].index = j
        # add numberOfPreviousActivities
        for Activity in currentActivitySequence.activities:
            for IndexFollowingActivity in Activity.indexFollowingActivities:
                currentActivitySequence.activities[IndexFollowingActivity].numberOfPreviousActivities += 1
    activitySequences.append(currentActivitySequence)

stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources

# compute decisions: each decision corresponds to a start of an activity in the local reference system (more than one decision can be taken at once)
for i in range(0, numberOfActivitiesInStateVector):
    decisions_indexActivity.append(i)

# record 10 times run data
for run in range(numberOfExperiments):
    # reset variable and lists at each run
    states = []
    actions = []
    importExistingNeuralNetworkModel = False
    neuralNetworkModelAlreadyExists = False

    # --------------------------------------------------------------RANDOM-----------------------------------------------------------------------------
    ####  GENERATE TRAINING DATA USING RANDOM DECISIONS (WITHOUT USING pool.map) ####
    print('######  RANDOM DECISION ON TRAIN ACTIVITY SEQUENCES  ######')
    runSimulation_inputs = []
    for i in range(numberOfFilesTrain):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "generateData"
        currentRunSimulation_input.randomDecisionProbability = 1
        currentRunSimulation_input.policyType = None
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        runSimulation_inputs.append(currentRunSimulation_input)

    pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

    runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
    # assign simulation results to activity sequences and append training data

    for i in range(numberOfFilesTrain):
        activitySequences[indexFilesTrain[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
        activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation = runSimulation_outputs[
            i].totalDurationStDev
        activitySequences[indexFilesTrain[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
        activitySequences[indexFilesTrain[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
        activitySequences[indexFilesTrain[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
        activitySequences[indexFilesTrain[i]].trivialDecisionPercentageMean = runSimulation_outputs[
            i].trivialDecisionPercentageMean

        for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
            states.append(list(currentStateActionPair.state))
            actions.append(list(currentStateActionPair.action))
        # correspondence best states and actions pairs --> len(states) = len(actions)
        # print('states:',len(states))
        # print('length of state',len(states[0]))
        # print('actions:',actions)

    ####  TRAIN MODEL USING TRAINING DATA  ####
    # look for existing model
    if importExistingNeuralNetworkModel:
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")

        else:
            neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
            # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]))
    else:
        neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
        # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]))

    states = np.array(states).reshape(-1, len(states[0]), 1)
    neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_step=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")

    ####  CREATE BENCHMARK WITH RANDOM DECISIONS ALSO WITH TEST ACTIVITY SEQUENCES  ####
    print('######  RANDOM DECISION ON TEST ACTIVITY SEQUENCES  ######')
    runSimulation_inputs = []
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 1
        currentRunSimulation_input.policyType = None
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        runSimulation_inputs.append(currentRunSimulation_input)

    pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

    runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
    # assign simulation results to activity sequences

    for i in range(numberOfFilesTest):
        activitySequences[indexFilesTest[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
        activitySequences[indexFilesTest[i]].totalDurationStandardDeviation = runSimulation_outputs[
            i].totalDurationStDev
        activitySequences[indexFilesTest[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
        activitySequences[indexFilesTest[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
        activitySequences[indexFilesTest[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
        activitySequences[indexFilesTest[i]].trivialDecisionPercentageMean = runSimulation_outputs[
            i].trivialDecisionPercentageMean

    # -----------------------------------------------------------------NN------------------------------------------------------------------------------
    ####  TEST NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ####
    # run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
    print('###### NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ######')
    for i in range(numberOfFilesTrain):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "neuralNetworkModel"
        currentRunSimulation_input.decisionTool = neuralNetworkModel
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[indexFilesTrain[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

    ####  TEST NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ####
    # run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
    print('###### NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ######')
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "neuralNetworkModel"
        currentRunSimulation_input.decisionTool = neuralNetworkModel
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[indexFilesTest[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

    # ---------------------------------------------------------Heuristic----------------------------------------------------------------------------
    ####  TEST HEURISTIC METHOD ON TRAIN ACTIVITY SEQUENCES  ####
    print('###### HEURISTIC METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
    runSimulation_inputs = []
    for i in range(numberOfFilesTrain):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "heuristic"
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        runSimulation_inputs.append(currentRunSimulation_input)

    pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

    runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
    # assign simulation results to activity sequences
    for i in range(numberOfFilesTrain):
        activitySequences[indexFilesTrain[i]].totalDurationWithHeuristic = currentRunSimulation_output.totalDurationMean

    print('###### HEURISTIC METHOD ON TEST ACTIVITY SEQUENCES  ######')
    runSimulation_inputs = []
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "heuristic"
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities

        runSimulation_inputs.append(currentRunSimulation_input)

    pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

    runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
    # assign simulation results to activity sequences
    for i in range(numberOfFilesTest):
        activitySequences[indexFilesTest[i]].totalDurationWithHeuristic = currentRunSimulation_output.totalDurationMean

    # ------------------------------------------------------EVALUATION-----------------------------------------------------------------------------
    ####  EVALUATION OF RESULTS OF TRAIN ACTIVITY SEQUENCES  ####
    sumTotalDurationRandomTrain = 0
    sumTotalDurationWithNeuralNetworkModelTrain = 0
    sumTotalDurationWithHeuristicTrain = 0

    for i in range(numberOfFilesTrain):
        sumTotalDurationRandomTrain += activitySequences[indexFilesTrain[i]].totalDurationMean
        sumTotalDurationWithNeuralNetworkModelTrain += activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
        sumTotalDurationWithHeuristicTrain += activitySequences[indexFilesTrain[i]].totalDurationWithHeuristic

    sumTotalDurationRandomTrainRecord.append(sumTotalDurationRandomTrain)
    sumTotalDurationWithNeuralNetworkModelTrainRecord.append(sumTotalDurationWithNeuralNetworkModelTrain)
    sumTotalDurationWithHeuristicTrainRecord.append(sumTotalDurationWithHeuristicTrain)

    ####  EVALUATION OF RESULTS OF TEST ACTIVITY SEQUENCES  ####
    sumTotalDurationRandomTest = 0
    sumTotalDurationWithNeuralNetworkModelTest = 0
    sumTotalDurationWithHeuristicTest = 0

    for i in range(numberOfFilesTest):
        sumTotalDurationRandomTest += activitySequences[indexFilesTest[i]].totalDurationMean
        sumTotalDurationWithNeuralNetworkModelTest += activitySequences[indexFilesTest[i]].totalDurationWithPolicy
        sumTotalDurationWithHeuristicTest += activitySequences[indexFilesTest[i]].totalDurationWithHeuristic

    run += 1

sumTotalDurationRandomTestRecord.append(sumTotalDurationRandomTest)
sumTotalDurationWithNeuralNetworkModelTestRecord.append(sumTotalDurationWithNeuralNetworkModelTest)
sumTotalDurationWithHeuristicTestRecord.append(sumTotalDurationWithHeuristicTest)

print("sumTotalDurationRandomTrain = " + str(sumTotalDurationRandomTrain))
print("sumTotalDurationWithNeuralNetworkModelTrain = " + str(sumTotalDurationWithNeuralNetworkModelTrain))
print("sumTotalDurationWithHeuristicTrain = " + str(sumTotalDurationWithHeuristicTrain))
print("sumTotalDurationRandomTest = " + str(sumTotalDurationRandomTest))
print("sumTotalDurationWithNeuralNetworkModelTest = " + str(sumTotalDurationWithNeuralNetworkModelTest))
print("sumTotalDurationWithHeuristicTest = " + str(sumTotalDurationWithHeuristicTest))

# compute computation time
t_end = time.time()
t_computation = t_end - t_start
print("t_computation = " + str(t_computation))

#     # write to excel
#     ws['A1'] = 'RandomTrain'
#     ws['B1'] = 'NNTrain'
#     ws['D1'] = 'RandomTest'
#     ws['F1'] = 'NNTest'
#     ws['H1'] = 'HeuristicTest'
#     ws['I1'] = 'time'
#     ws.cell(row=run + 2, column=4).value = sumTotalDurationRandomTestRecord[0]
#     ws.cell(row=run + 2, column=6).value = sumTotalDurationWithNeuralNetworkModelTestRecord[0]
#     ws.cell(row=run + 2, column=1).value = sumTotalDurationRandomTrainRecord[0]
#     ws.cell(row=run + 2, column=2).value = sumTotalDurationWithNeuralNetworkModelTrainRecord[0]
#     ws.cell(row=run + 2, column=8).value = sumTotalDurationWithHeuristicTestRecord[0]
#     ws.cell(row=run + 2, column=3).value = 1
#     ws.cell(row=run + 2, column=5).value = 2
#     ws.cell(row=run + 2, column=7).value = 3
#     ws.cell(row=2, column=9).value = t_computation
#
#     #change column width and height
#     ws.column_dimensions['A'].width = 15.0
#     ws.column_dimensions['D'].width = 15.0
#     ws.column_dimensions['H'].width = 15.0
#
#     # alignment
#     align = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     ws['A1'].alignment = align
#     ws['B1'].alignment = align
#     ws['D1'].alignment = align
#     ws['F1'].alignment = align
#     ws['H1'].alignment = align
#     ws['I1'].alignment = align
#
#
# wb.save(relativePath + "/database_480/1000times3hidden.xlsx")

# write ouput to excel
#

